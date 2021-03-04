from openpyxl import load_workbook
from openpyxl import Workbook
import argparse

App_list = ["Phone", "Messages/Messaging", "Settings", "Contacts", "Camera", "Album/Gallery", "Music", "FM Radio",
            "Email", "Clock", "File Manager", "Voice Note/Voice Recorder", "Calculator", "Browser", "Play Store/應用商店",
            "Youtube", "Maps", "Duo", "Facebook"]

worse_list = []


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--file_name", required=True, help='result file name')
    parser.add_argument("-l", "--line_number", required=True, help='line_number from 0')
    parser.add_argument("-b", "--current_build", required=True, help='build args.line_number')
    # parser.add_argument('-v', "--verbose", action='store_true', default=False,
    #                     help='output verbose information if specified')
    return parser.parse_args()


def _main():
    args = parse_args()
    wb = load_workbook(args.file_name, data_only=True)
    test_data = wb["Case1_TestData"]
    book = Workbook()
    sheet = book.create_sheet(args.current_build)
    total = 0
    curse = 2
    sheet_no = []
    vs_list = []
    for i in range(ord("B"), ord("B") + int(args.line_number) + 1):
        vs = "V0." + test_data["C" + str(curse)].value.split("_")[-1]
        curse += 1
        vs_list.append(vs)
        sheet_no.append(chr(i) + "1")
    for k, v in dict(zip(sheet_no, vs_list)).items():
        sheet[k] = v

    sheet[chr(ord("B") + int(args.line_number) + 1) + "1"] = "Gap"
    sheet[chr(ord("B") + int(args.line_number) + 2) + "1"] = "Percentage"

    for i in range(2, 676, 15):
        app_name = test_data["B" + str(i)].value
        version_list = []
        avg_list = []
        for j in range(int(args.line_number) + 1):
            version = test_data["C" + str(j + 2)].value.split("_")[-1]
            version_list.append("V0." + version)
            avg_list.append(test_data["E" + str(i + j)].value)
        if app_name in App_list:
            if avg_list[-1] > avg_list[0]:
                total += 1
                sheet["A" + str(total + 1)] = app_name
                for k in range(len(avg_list)):
                    sheet[chr(ord("B") + k) + str(total + 1)] = avg_list[k]
                worse_list.append(app_name)
                gap = avg_list[-1] - avg_list[0]
                percentage = gap / avg_list[0]
                sheet[chr(ord("B") + int(args.line_number) + 1) + str(total + 1)] = round(gap, 2)
                sheet[chr(ord("B") + int(args.line_number) + 2) + str(total + 1)] = "{:.2%}".format(percentage)
                book.save(r"./loadingTestResult.xlsx")

    print("--------------------------------test result-----------------------------------------")
    print(f"{total} of {args.current_build} is worse than V0.240")
    print("worse list:")
    print(",".join(worse_list))
    print("------------------------------------------------------------------------------------")


if __name__ == '__main__':
    _main()
